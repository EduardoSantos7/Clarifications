import os
import pandas as pd
from openpyxl import load_workbook

class File():

    # For the moment just data is not allowed
    def __init__(self, data=None,  to_read="", to_save="", 
                    allowed=[], read_options={}, save_options={}):

        self.data = data or pd.DataFrame()
        self.rows = 0
        self.columns = 0
        self.allowed_extensions = allowed or ['xlsx', 'xls', 'csv']
        self.to_read = to_read
        self.to_save = to_save

        if self.to_read:
            self.read(self.to_read, **read_options)

    def read(self, paths, **read_options):

        # If input is a string put inside a list.
        if type(paths) == str:
            paths = [paths]
        
        # In case the function is called outside the __init__
        # set the 'to_read' variable.
        if not self.to_read and paths:
            self.to_read = paths
        
        for path in paths:
            file_name = path.split('/')[-1]
            path = path.replace(file_name, '')
            extension = file_name.split('.')
            
            # If a path is provided change the directory.
            if path:
                os.chdir(path)
            
            if extension[-1] in self.allowed_extensions:
                if extension[-1] == "xlsx" or extension[-1] == 'xls':
                    self.data = pd.read_excel(file_name)
                if extension[-1] == "csv":
                    self.data = pd.read_csv(file_name, **read_options)
                
                # Refresh row and columns count
                self.rows = len(self.data.index)
                self.columns = len(self.data.columns)
            else:
                print(f"Extension {extension[-1]} is not allowed")

    def save(self, path="", **save_options):

        path = path or self.to_save
        if not self.to_save and path:
            self.to_save = path

        if path:

            file_name = path.split('/')[-1]
            path = path.replace(file_name, '')
            extension = file_name.split('.')
            
            # If a path is provided change the directory.
            if path:
                os.chdir(path)

            extension = file_name.split('.')

            if len(extension) != 2:
                print(f"The name: {file_name} is incorrect")
                return
            
            if extension[1] in self.allowed_extensions:
                if extension[1] == "xlsx" or extension[1] == 'xls':
                    self.data.to_excel(file_name, **save_options)
                if extension[1] == "csv":
                    self.data.to_csv(file_name, **save_options)
                print("after print", extension[1])
            else:
                print(f"Extension {extension[1]} is not allowed")

    def append_df_to_excel(self, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
        """
        Append a DataFrame [df] to existing Excel file [self.to_save]
        into [sheet_name] Sheet.
        If [self.to_save] doesn't exist, then this function will create it.

        Parameters:
        filename : File path or existing ExcelWriter
                    (Example: '/path/to/file.xlsx')
        df : dataframe to save to workbook
        sheet_name : Name of sheet which will contain DataFrame.
                    (default: 'Sheet1')
        startrow : upper left cell row to dump data frame.
                    Per default (startrow=None) calculate the last row
                    in the existing DF and write to the next row...
        truncate_sheet : truncate (remove and recreate) [sheet_name]
                        before writing DataFrame to Excel file
        to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                            [can be dictionary]

        Returns: None

        Use:
            append_df_to_excel(df, **{'header':None}) # Use to avoid new df headers
        """

        # ignore [engine] parameter if it was passed
        if 'engine' in to_excel_kwargs:
            to_excel_kwargs.pop('engine')

        writer = pd.ExcelWriter(self.to_save, engine='openpyxl')

        try:
            # try to open an existing workbook
            writer.book = load_workbook(self.to_save)

            # get the last row in the existing Excel sheet
            # if it was not specified explicitly
            if startrow is None and sheet_name in writer.book.sheetnames:
                startrow = writer.book[sheet_name].max_row

            # truncate sheet
            if truncate_sheet and sheet_name in writer.book.sheetnames:
                # index of [sheet_name] sheet
                idx = writer.book.sheetnames.index(sheet_name)
                # remove [sheet_name]
                writer.book.remove(writer.book.worksheets[idx])
                # create an empty sheet [sheet_name] using old index
                writer.book.create_sheet(sheet_name, idx)

            # copy existing sheets
            writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
        except FileNotFoundError:
            # file does not exist yet, we will create it
            pass

        if startrow is None:
            startrow = 0

        # write out the new sheet
        df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

        # save the workbook
        writer.save()

    def delete_duplicates(self, inplace=True, keep='first', subset=None):
        """ Remove duplicates"""
        if not self.data.empty and inplace:
            self.data.drop_duplicates(inplace=inplace, keep=keep, subset=subset)
        elif not self.data.empty:
            return self.data.drop_duplicates(inplace=inplace, keep=keep, subset=subset)

    def remove_rows_by_columns(self, colum_name, value):
        """ Remove the rows in which the [colum_name] 
        has the indicated [value]"""
        
        self.data = self.data[self.data[colum_name] != ( value)]
        self.update_size()
    
    def concat(self, dfs, **concat_options):
        
        if type(dfs) == list:
            print("here")
            self.data = pd.concat([self.data] +  dfs, **concat_options)
        elif not dfs.empty:
            print("here2")
            self.data = pd.concat([self.data, dfs], **concat_options)

        self.update_size()

    def update_size(self):
        
        # Refresh row and columns count
        self.rows = len(self.data.index)
        self.columns = len(self.data.columns)

